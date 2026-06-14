from collections import defaultdict
from datetime import datetime
from pathlib import Path

from django.utils.text import slugify
from openpyxl import load_workbook


MONTHS = [
    "September 2025",
    "October 2025",
    "November 2025",
    "December 2025",
    "January 2026",
    "February 2026",
    "March 2026",
    "April 2026",
    "May 2026",
]

COMBINED_HEADERS = [
    "Month",
    "Rank",
    "Title",
    "Primary_Artist",
    "Featured_Artists",
    "Combined_Points_Raw",
    "Display_Points",
    "Platforms",
    "Platforms_Max",
    "Weeks",
    "Release_Year",
    "Confidence",
]

PLATFORM_HEADERS = ["Month", "Platform", "Rank", "Title", "Artist", "Points", "Weeks"]

PLATFORM_DATA = [
    ("Apple Music", "apple-music", "#FC3C44", 200, 101),
    ("Audiomack", "audiomack", "#F68B1F", 200, 101),
    ("Boomplay", "boomplay", "#00FFFF", 100, 101),
    ("Spotify", "spotify", "#1DB954", 50, 101),
    ("YouTube", "youtube", "#FF0000", 100, 101),
    ("Shazam", "shazam", "#0088FF", 100, 101),
]


def _sheet_records(workbook, sheet_name, expected_headers):
    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    if headers != expected_headers:
        raise ValueError(f"{sheet_name} headers do not match the documented workbook schema")
    return [dict(zip(headers, row)) for row in rows if any(value is not None for value in row)]


def load_master_workbook(workbook_path):
    workbook_path = Path(workbook_path)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    required = {
        "README",
        "Singles_Combined",
        "Albums_Combined",
        "Singles_Platforms",
        "Albums_Platforms",
    }
    if set(workbook.sheetnames) != required:
        raise ValueError("The master workbook must contain exactly the five documented sheets")

    data = {
        "singles": {
            "combined": _sheet_records(workbook, "Singles_Combined", COMBINED_HEADERS),
            "platforms": _sheet_records(workbook, "Singles_Platforms", PLATFORM_HEADERS),
        },
        "albums": {
            "combined": _sheet_records(workbook, "Albums_Combined", COMBINED_HEADERS),
            "platforms": _sheet_records(workbook, "Albums_Platforms", PLATFORM_HEADERS),
        },
    }
    workbook.close()
    validate_master_data(data)
    return data


def validate_master_data(data):
    combined_total = 0
    for chart_type in ("singles", "albums"):
        rows = data[chart_type]["combined"]
        combined_total += len(rows)
        grouped = defaultdict(list)
        for row in rows:
            grouped[row["Month"]].append(row)

        if list(grouped) != MONTHS:
            raise ValueError(f"{chart_type} Combined months are missing or out of order")

        expected_max = 6 if chart_type == "singles" else 2
        for month in MONTHS:
            month_rows = grouped[month]
            ranks = [int(row["Rank"]) for row in month_rows]
            if len(month_rows) != 50 or ranks != list(range(1, 51)):
                raise ValueError(f"{chart_type} {month} must contain ranks 1 through 50 exactly once")
            for row in month_rows:
                rank = int(row["Rank"])
                if int(row["Display_Points"]) != 51 - rank:
                    raise ValueError(f"Invalid Display_Points at {chart_type} {month} rank {rank}")
                if int(row["Platforms_Max"]) != expected_max:
                    raise ValueError(f"Invalid Platforms_Max at {chart_type} {month} rank {rank}")

    if combined_total != 900:
        raise ValueError("The two Combined sheets must contain exactly 900 rows")


def _month_parts(label):
    parsed = datetime.strptime(label, "%B %Y")
    return parsed.year, parsed.month


def _safe_slug(name, used):
    base = slugify(name)[:45] or "unknown"
    candidate = base
    suffix = 2
    while candidate in used and used[candidate] != name:
        candidate = f"{base[:40]}-{suffix}"
        suffix += 1
    used[candidate] = name
    return candidate


def import_master_workbook(app_registry, workbook_path, clear=True, write_line=None):
    data = load_master_workbook(workbook_path)
    write_line = write_line or (lambda _message: None)

    Platform = app_registry.get_model("charts", "Platform")
    Artist = app_registry.get_model("charts", "Artist")
    Release = app_registry.get_model("charts", "Release")
    MonthlyChart = app_registry.get_model("charts", "MonthlyChart")
    MonthlyChartEntry = app_registry.get_model("charts", "MonthlyChartEntry")
    WeeklyUpload = app_registry.get_model("charts", "WeeklyUpload")
    Certification = app_registry.get_model("charts", "Certification")

    for name, slug, color, chart_size, points_base in PLATFORM_DATA:
        Platform.objects.update_or_create(
            name=name,
            defaults={
                "slug": slug,
                "color": color,
                "chart_size": chart_size,
                "points_base": points_base,
                "active": True,
            },
        )
    platforms = {platform.name: platform for platform in Platform.objects.all()}

    source_artist_names = {
        str(row["Primary_Artist"]).strip()
        for chart_type in data.values()
        for row in chart_type["combined"]
    }
    source_artist_names.update(
        str(row["Artist"]).strip()
        for chart_type in data.values()
        for row in chart_type["platforms"]
    )

    if clear:
        write_line("Clearing existing singles and albums chart data...")
        Certification.objects.all().delete()
        MonthlyChartEntry.objects.all().delete()
        MonthlyChart.objects.all().delete()
        WeeklyUpload.objects.all().delete()
        Release.objects.all().delete()
        Artist.objects.exclude(name__in=source_artist_names).delete()

    artist_cache = {artist.name: artist for artist in Artist.objects.all()}
    used_slugs = {artist.slug: artist.name for artist in artist_cache.values()}
    release_cache = {}
    chart_cache = {}

    def get_artist(name):
        name = str(name).strip()
        artist = artist_cache.get(name)
        if artist is None:
            artist = Artist.objects.create(name=name, slug=_safe_slug(name, used_slugs))
            artist_cache[name] = artist
        return artist

    def get_release(title, artist_name, chart_type):
        title = str(title).strip()
        artist = get_artist(artist_name)
        key = (chart_type, title.casefold(), artist.pk)
        release = release_cache.get(key)
        if release is None:
            release = Release.objects.create(
                title=title,
                artist=artist,
                chart_type=chart_type,
                canonical_title=title.casefold(),
            )
            release_cache[key] = release
        return release

    def get_chart(chart_type, month_label):
        key = (chart_type, month_label)
        chart = chart_cache.get(key)
        if chart is None:
            year, month_number = _month_parts(month_label)
            chart = MonthlyChart.objects.create(
                year=year,
                month=month_number,
                chart_type=chart_type,
                label=month_label,
                is_published=True,
            )
            chart_cache[key] = chart
        return chart

    combined_count = 0
    platform_count = 0

    for chart_type in ("singles", "albums"):
        combined_groups = defaultdict(list)
        for row in data[chart_type]["combined"]:
            combined_groups[row["Month"]].append(row)

        previous_rank = {}
        peak_rank = {}
        for month_label in MONTHS:
            chart = get_chart(chart_type, month_label)
            rows = combined_groups[month_label]
            current_best = {}
            for row in rows:
                key = (str(row["Title"]).strip().casefold(), str(row["Primary_Artist"]).strip().casefold())
                current_best[key] = min(current_best.get(key, 999), int(row["Rank"]))

            entries = []
            for row in rows:
                key = (str(row["Title"]).strip().casefold(), str(row["Primary_Artist"]).strip().casefold())
                rank = int(row["Rank"])
                release = get_release(row["Title"], row["Primary_Artist"], chart_type)
                entries.append(
                    MonthlyChartEntry(
                        chart=chart,
                        platform=None,
                        release=release,
                        rank=rank,
                        total_points=int(row["Display_Points"]),
                        raw_total_points=int(row["Combined_Points_Raw"]),
                        weeks_on_chart=int(row["Weeks"]),
                        platform_count=int(row["Platforms"]),
                        platform_max=int(row["Platforms_Max"]),
                        featured_artists=str(row["Featured_Artists"] or "").strip(),
                        release_year=int(row["Release_Year"]) if row["Release_Year"] is not None else None,
                        confidence=str(row["Confidence"] or "").strip(),
                        peak_rank=min(peak_rank.get(key, 999), current_best[key]),
                        prev_rank=previous_rank.get(key),
                    )
                )
            MonthlyChartEntry.objects.bulk_create(entries, batch_size=500)
            combined_count += len(entries)
            for key, rank in current_best.items():
                previous_rank[key] = rank
                peak_rank[key] = min(peak_rank.get(key, 999), rank)

        platform_groups = defaultdict(list)
        for row in data[chart_type]["platforms"]:
            platform_groups[(row["Month"], row["Platform"])].append(row)

        platform_previous = {}
        platform_peak = {}
        platform_names = ["Apple Music", "Audiomack"] if chart_type == "albums" else [item[0] for item in PLATFORM_DATA]
        for month_label in MONTHS:
            chart = get_chart(chart_type, month_label)
            for platform_name in platform_names:
                platform = platforms[platform_name]
                rows = platform_groups[(month_label, platform_name)]
                entries = []
                for row in rows:
                    key = (
                        platform_name,
                        str(row["Title"]).strip().casefold(),
                        str(row["Artist"]).strip().casefold(),
                    )
                    rank = int(row["Rank"])
                    release = get_release(row["Title"], row["Artist"], chart_type)
                    entries.append(
                        MonthlyChartEntry(
                            chart=chart,
                            platform=platform,
                            release=release,
                            rank=rank,
                            total_points=int(row["Points"]),
                            raw_total_points=None,
                            weeks_on_chart=int(row["Weeks"]),
                            platform_count=1,
                            platform_max=1,
                            featured_artists="",
                            release_year=None,
                            confidence="",
                            peak_rank=min(platform_peak.get(key, 999), rank),
                            prev_rank=platform_previous.get(key),
                        )
                    )
                    platform_previous[key] = rank
                    platform_peak[key] = min(platform_peak.get(key, 999), rank)
                MonthlyChartEntry.objects.bulk_create(entries, batch_size=500)
                platform_count += len(entries)

    write_line(f"Imported {combined_count} Combined rows and {platform_count} platform rows")
    return {
        "combined_rows": combined_count,
        "platform_rows": platform_count,
        "total_rows": combined_count + platform_count,
        "months": MONTHS,
    }
