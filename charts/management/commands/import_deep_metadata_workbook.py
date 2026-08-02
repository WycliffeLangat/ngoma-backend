"""Import the deep metadata research workbook into public Release/Artist data.

The workbook is row-oriented and does not carry database IDs, so records are
matched by chart type, title/canonical title, and credited artist text. Exact
release dates are imported as dates; month/year/upper-bound rows are not forced
into the exact DateField.
"""
import json
import os
import re
from collections import Counter, defaultdict
from datetime import date, datetime

import openpyxl
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone

from charts.artist_credits import release_credit_payload
from charts.cms_utils import bump_public_revision
from charts.models import Artist, Release


RELEASE_SHEET = "Merged Releases"
ARTIST_SHEET = "Artists"

PLACEHOLDER_VALUES = {
    "",
    "n/a",
    "na",
    "none",
    "null",
    "not available",
    "not publicly available",
    "none - availability documented",
    "none — availability documented",
    "no additional notes.",
    "complete",
}

RELEASE_FIELDS = [
    "songwriters",
    "producers",
    "release_year",
    "release_date",
    "isrc",
    "upc",
    "number_of_tracks",
    "genre",
    "label",
    "distributor",
    "apple_music_url",
]

ARTIST_FIELDS = [
    "country",
    "country_code",
    "city_region",
    "genre",
]

SKIP_VALUE = object()


def clean_text(value):
    if value is None:
        return ""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if float(value).is_integer():
            return str(int(value))
        return str(value).strip()
    return str(value).strip()


def meaningful(value):
    text = clean_text(value)
    if text.casefold() in PLACEHOLDER_VALUES:
        return ""
    return text


def normalized(value):
    text = clean_text(value).casefold()
    text = text.replace("&", " and ")
    text = re.sub(r"\b(feat|ft|featuring)\.?\b", " and ", text)
    return re.sub(r"[\W_]+", " ", text, flags=re.U).strip()


def title_variants(value):
    text = clean_text(value)
    variants = {normalized(text)}
    variants.add(normalized(re.sub(r"\s*\((?:feat|ft|featuring)\.?.*?\)", "", text, flags=re.I)))
    variants.add(normalized(text.replace("...", "")))
    return {variant for variant in variants if variant}


def name_parts(value):
    text = clean_text(value)
    if not text:
        return set()
    pieces = re.split(r"\s*(?:;|,|&|\band\b|\bx\b|\bfeat\.?\b|\bft\.?\b|\bfeaturing\b)\s*", text, flags=re.I)
    return {normalized(piece) for piece in pieces if normalized(piece)}


def parse_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean_text(value)
    match = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", text)
    if not match:
        return None
    try:
        return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
    except ValueError:
        return None


def parse_year(value):
    text = clean_text(value)
    match = re.match(r"^(\d{4})", text)
    if not match:
        return None
    year = int(match.group(1))
    return year if 1900 <= year <= 2100 else None


def parse_int(value):
    text = meaningful(value)
    if not text:
        return None
    match = re.match(r"^\d+$", text)
    return int(text) if match else None


def is_url(value):
    return clean_text(value).startswith(("http://", "https://"))


def model_safe_value(model_class, field_name, value):
    if value is None:
        return value
    field = model_class._meta.get_field(field_name)
    max_length = getattr(field, "max_length", None)
    if max_length and len(clean_text(value)) > max_length:
        return SKIP_VALUE
    return value


def rows_from_sheet(workbook, sheet_name, header_row=1):
    if sheet_name not in workbook.sheetnames:
        raise CommandError(f"Workbook is missing required sheet: {sheet_name}")
    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    for _ in range(header_row - 1):
        next(rows, None)
    headers = [clean_text(value) for value in next(rows, [])]
    index = {header: offset for offset, header in enumerate(headers) if header}
    for raw in rows:
        if not any(value is not None and clean_text(value) for value in raw):
            continue
        yield {header: raw[offset] if offset < len(raw) else None for header, offset in index.items()}


def release_chart_type(row):
    chart = normalized(row.get("Chart"))
    if chart.startswith("album"):
        return "albums"
    return "singles"


def release_title_keys(row):
    keys = set()
    for value in [row.get("Canonical title"), row.get("Title")]:
        keys.update(title_variants(value))
    return keys


def release_artist_keys(row):
    values = [
        row.get("Canonical artists"),
        row.get("Artists"),
    ]
    return {normalized(value) for value in values if normalized(value)}


def release_artist_parts(row):
    parts = set()
    for value in [row.get("Canonical artists"), row.get("Artists")]:
        parts.update(name_parts(value))
    return parts


def release_index():
    releases = (
        Release.objects.select_related("artist")
        .prefetch_related("artist_credits__artist")
        .all()
    )
    by_chart_title = defaultdict(list)
    metadata = {}
    for release in releases:
        credit = release_credit_payload(release)
        title_keys = title_variants(release.title) | title_variants(release.canonical_title)
        artist_texts = {
            normalized(release.artist.name),
            normalized(release.artist.display_name),
            normalized(credit.get("artist_credit")),
            normalized(credit.get("primary_artist_credit")),
        }
        artist_part_set = name_parts(credit.get("artist_credit")) | name_parts(credit.get("primary_artist_credit")) | name_parts(release.artist.name)
        metadata[release.id] = {
            "release": release,
            "title_keys": {key for key in title_keys if key},
            "artist_keys": {key for key in artist_texts if key},
            "artist_parts": artist_part_set,
        }
        for key in metadata[release.id]["title_keys"]:
            by_chart_title[(release.chart_type, key)].append(release.id)
    return by_chart_title, metadata


def match_release(row, by_chart_title, metadata):
    chart_type = release_chart_type(row)
    title_keys = release_title_keys(row)
    artist_keys = release_artist_keys(row)
    artist_parts = release_artist_parts(row)
    candidate_ids = []
    seen = set()
    for title_key in title_keys:
        for release_id in by_chart_title.get((chart_type, title_key), []):
            if release_id not in seen:
                seen.add(release_id)
                candidate_ids.append(release_id)
    if not candidate_ids:
        return None, "unmatched"

    scored = []
    for release_id in candidate_ids:
        data = metadata[release_id]
        full_artist_match = bool(artist_keys & data["artist_keys"])
        row_parts = artist_parts
        release_parts = data["artist_parts"]
        part_overlap = bool(row_parts & release_parts)
        set_match = bool(row_parts and release_parts and (row_parts <= release_parts or release_parts <= row_parts))
        score = (3 if full_artist_match else 0) + (2 if set_match else 0) + (1 if part_overlap else 0)
        scored.append((score, release_id))

    scored.sort(reverse=True)
    if scored[0][0] <= 0:
        if len(candidate_ids) == 1:
            return metadata[candidate_ids[0]]["release"], "title-only"
        return None, "ambiguous-title"
    if len(scored) > 1 and scored[0][0] == scored[1][0]:
        return None, "ambiguous"
    return metadata[scored[0][1]]["release"], "matched"


def release_values(row, clear_non_exact_dates=True):
    values = {}
    precision = normalized(row.get("Precision"))
    exact_date = parse_date(row.get("Release date")) if precision == "exact" else None
    year = parse_year(row.get("Year"))

    if exact_date:
        values["release_date"] = exact_date
        values["release_year"] = exact_date.year
    elif precision in {"month", "year"} and year:
        values["release_year"] = year
        if clear_non_exact_dates:
            values["release_date"] = None
    elif precision == "upper bound" and clear_non_exact_dates:
        values["release_date"] = None

    field_map = {
        "isrc": "ISRC",
        "upc": "UPC",
        "label": "Label",
        "distributor": "Distributor",
        "producers": "Producer(s)",
        "songwriters": "Songwriter(s)",
    }
    for field, column in field_map.items():
        value = meaningful(row.get(column))
        if value:
            values[field] = value

    track_count = parse_int(row.get("Track count"))
    if track_count is not None:
        values["number_of_tracks"] = track_count

    genre = meaningful(row.get("Genre (web)")) or meaningful(row.get("Genre (Apple)"))
    if genre:
        values["genre"] = genre

    apple_url = meaningful(row.get("Apple source URL"))
    if apple_url and is_url(apple_url):
        values["apple_music_url"] = apple_url

    return values


def artist_index():
    artists = Artist.objects.all()
    by_key = defaultdict(list)
    for artist in artists:
        values = [artist.name, artist.display_name, *(artist.aliases or [])]
        for value in values:
            key = normalized(value)
            if key:
                by_key[key].append(artist)
    return by_key


def match_artist(row, by_key):
    key = normalized(row.get("Artist"))
    if not key:
        return None, "blank"
    matches = by_key.get(key, [])
    if len(matches) == 1:
        return matches[0], "matched"
    if len(matches) > 1:
        return None, "ambiguous"
    return None, "unmatched"


def artist_values(row):
    values = {}
    country = meaningful(row.get("Country"))
    code = meaningful(row.get("Code")).upper()
    city = meaningful(row.get("City / Region"))
    genre = meaningful(row.get("Genre"))

    if country:
        values["country"] = country
    if re.match(r"^[A-Z]{2}$", code):
        values["country_code"] = code
    if city:
        values["city_region"] = city
    if genre:
        values["genre"] = genre
    return values


class Command(BaseCommand):
    help = "Import release and artist metadata from ngoma_metadata_release_dates_deep_research.xlsx."

    def add_arguments(self, parser):
        parser.add_argument("--path", required=True, help="Path to ngoma_metadata_release_dates_deep_research.xlsx")
        parser.add_argument("--apply", action="store_true", help="Write changes. Without this flag this is a dry run.")
        parser.add_argument("--overwrite", action="store_true", help="Replace existing values with meaningful workbook values.")
        parser.add_argument(
            "--keep-non-exact-dates",
            action="store_true",
            help="Do not clear release_date for rows whose workbook precision is month/year/upper bound.",
        )
        parser.add_argument("--backup-path", default="", help="Optional JSON backup path for records touched during --apply.")
        parser.add_argument("--sample-limit", type=int, default=25)

    def handle(self, *args, **options):
        path = options["path"]
        if not os.path.exists(path):
            raise CommandError(f"Workbook not found: {path}")

        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        release_rows = list(rows_from_sheet(workbook, RELEASE_SHEET))
        artist_rows = list(rows_from_sheet(workbook, ARTIST_SHEET))

        release_changes, release_stats = self._plan_releases(
            release_rows,
            overwrite=options["overwrite"],
            clear_non_exact_dates=not options["keep_non_exact_dates"],
        )
        artist_changes, artist_stats = self._plan_artists(artist_rows, overwrite=options["overwrite"])

        self._report("Releases", release_changes, release_stats, options["sample_limit"])
        self._report("Artists", artist_changes, artist_stats, options["sample_limit"])

        if not options["apply"]:
            self.stdout.write(self.style.WARNING("\nDRY RUN -- no database changes written. Re-run with --apply to save."))
            return

        backup_path = options["backup_path"] or os.path.join(
            "scripts",
            f"metadata_workbook_import_backup_{timezone.now().strftime('%Y%m%d_%H%M%S')}.json",
        )
        self._write_backup(backup_path, release_changes, artist_changes)

        with transaction.atomic():
            updated_at = timezone.now()
            for change in release_changes:
                obj = change["obj"]
                for field, value in change["values"].items():
                    setattr(obj, field, value)
                obj.updated_at = updated_at
            for change in artist_changes:
                obj = change["obj"]
                for field, value in change["values"].items():
                    setattr(obj, field, value)
                obj.updated_at = updated_at
            release_update_fields = sorted({field for change in release_changes for field in change["values"]}) + ["updated_at"]
            artist_update_fields = sorted({field for change in artist_changes for field in change["values"]}) + ["updated_at"]
            if release_changes:
                Release.objects.bulk_update(
                    [change["obj"] for change in release_changes],
                    release_update_fields,
                    batch_size=200,
                )
            if artist_changes:
                Artist.objects.bulk_update(
                    [change["obj"] for change in artist_changes],
                    artist_update_fields,
                    batch_size=200,
                )
            bump_public_revision()

        self.stdout.write(self.style.SUCCESS(
            f"\nAPPLIED: {len(release_changes)} release(s), {len(artist_changes)} artist(s) updated."
        ))
        self.stdout.write(f"Backup: {backup_path}")

    def _plan_releases(self, rows, overwrite=False, clear_non_exact_dates=True):
        by_chart_title, metadata = release_index()
        changes = []
        stats = Counter()
        field_counts = Counter()
        for row in rows:
            release, status = match_release(row, by_chart_title, metadata)
            stats[status] += 1
            if not release:
                continue
            values = release_values(row, clear_non_exact_dates=clear_non_exact_dates)
            selected = {}
            for field, raw_value in values.items():
                value = model_safe_value(Release, field, raw_value)
                if value is SKIP_VALUE:
                    stats[f"{field}_too_long"] += 1
                    continue
                current = getattr(release, field)
                if field == "release_date":
                    current_value = current
                else:
                    current_value = clean_text(current)
                if value is None:
                    if current is not None and clear_non_exact_dates:
                        selected[field] = None
                elif field == "number_of_tracks":
                    if current != value and (overwrite or current in (None, 0)):
                        selected[field] = value
                elif field == "release_year":
                    if current != value and (overwrite or current is None):
                        selected[field] = value
                elif field == "release_date":
                    if current_value != value and (overwrite or current is None):
                        selected[field] = value
                else:
                    if current_value != clean_text(value) and (overwrite or not current_value):
                        selected[field] = value
            if selected:
                for field in selected:
                    field_counts[field] += 1
                changes.append({
                    "obj": release,
                    "values": selected,
                    "label": f"#{release.id} {release.title} - {release.artist.name}",
                    "workbook": {
                        "chart": clean_text(row.get("Chart")),
                        "title": clean_text(row.get("Title")),
                        "artists": clean_text(row.get("Artists")),
                        "precision": clean_text(row.get("Precision")),
                    },
                })
        stats["fields"] = dict(field_counts)
        return changes, stats

    def _plan_artists(self, rows, overwrite=False):
        by_key = artist_index()
        changes = []
        stats = Counter()
        field_counts = Counter()
        for row in rows:
            artist, status = match_artist(row, by_key)
            stats[status] += 1
            if not artist:
                continue
            values = artist_values(row)
            selected = {}
            for field, raw_value in values.items():
                value = model_safe_value(Artist, field, raw_value)
                if value is SKIP_VALUE:
                    stats[f"{field}_too_long"] += 1
                    continue
                current = clean_text(getattr(artist, field))
                if current != clean_text(value) and (overwrite or not current):
                    selected[field] = value
            if selected:
                for field in selected:
                    field_counts[field] += 1
                changes.append({
                    "obj": artist,
                    "values": selected,
                    "label": f"#{artist.id} {artist.name}",
                    "workbook": {
                        "artist": clean_text(row.get("Artist")),
                        "status": clean_text(row.get("Status")),
                    },
                })
        stats["fields"] = dict(field_counts)
        return changes, stats

    def _report(self, title, changes, stats, limit):
        fields = stats.pop("fields", {})
        self.stdout.write(self.style.MIGRATE_HEADING(f"\n{title}: {len(changes)} record(s) to update"))
        self.stdout.write(f"  Match stats: {json.dumps(dict(stats), sort_keys=True)}")
        self.stdout.write(f"  Field updates: {json.dumps(fields, sort_keys=True)}")
        for change in changes[:limit]:
            fields_text = ", ".join(f"{field}={value!s}" for field, value in change["values"].items())
            self.stdout.write(f"  {change['label']}: {fields_text}")
        if len(changes) > limit:
            self.stdout.write(f"  ...and {len(changes) - limit} more.")

    def _write_backup(self, path, release_changes, artist_changes):
        payload = {
            "generated_at": timezone.now().isoformat(),
            "releases": [
                {
                    "id": change["obj"].id,
                    "label": change["label"],
                    "current": {field: clean_text(getattr(change["obj"], field)) for field in change["values"]},
                    "incoming": {field: clean_text(value) for field, value in change["values"].items()},
                    "workbook": change["workbook"],
                }
                for change in release_changes
            ],
            "artists": [
                {
                    "id": change["obj"].id,
                    "label": change["label"],
                    "current": {field: clean_text(getattr(change["obj"], field)) for field in change["values"]},
                    "incoming": {field: clean_text(value) for field, value in change["values"].items()},
                    "workbook": change["workbook"],
                }
                for change in artist_changes
            ],
        }
        os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
        with open(path, "w", encoding="utf-8") as handle:
            json.dump(payload, handle, indent=2, ensure_ascii=False)
